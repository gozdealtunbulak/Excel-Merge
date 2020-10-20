import openpyxl
import xlrd
from openpyxl import Workbook, load_workbook
import pandas as pd
import time

name = 'test2'
extension = 'xlsx'
path = r'C:\Users\GOZDE\Desktop\excel'
f_name = name + '.' + extension
wb = xlrd.open_workbook('{}\{}'.format(path, f_name))
wb1 = openpyxl.load_workbook('{}\{}'.format(path, f_name))


def check_list(s_num, num):
    """
    :param s_num:the sheet index
    :type s_num int
    :param num: the row number
    :type num int
    :return: this function returns a list of the desired row of the chosen sheet
    """
    sheet = wb.sheet_by_index(s_num)
    list_check = []
    row_check = sheet.row_values(num)
    list_check.extend(row_check)
    return list_check


def add_new_sheet(sheet_name, index):
    """
    :param index: index of the sheet
    :type index int
    :param sheet_name: name of the sheet to be created
    :type sheet_name str
    :return:create the sheet
    """
    wb1.create_sheet(sheet_name, index=index)
    wb1.save('{}\{}'.format(path, f_name))


def remove_sheet(s_name):
    """"
    :param s_name of the sheet
    :type s_name str
    :return: removes the desired sheet
    """
    wb3 = openpyxl.load_workbook('{}\{}'.format(path, f_name))
    sheet = wb3[s_name]
    wb3.remove(sheet)
    wb3.save('{}\{}'.format(path, f_name))


def read_row_num(s_name):
    """
    :param s_name: name of the sheet
    :type s_name str
    :return:reads the row number of the sheet desired
    """
    wb2 = xlrd.open_workbook('{}\{}'.format(path, f_name))
    wk = wb2.sheet_by_name(s_name)
    existing_row = wk.nrows
    return existing_row


def add_data(s_name, data):
    """
    :param data: data to be added
    :type s_name str
    :param s_name: name of the sheet
    :type data list
    :return add the desired data to the desired sheet from the last row
    """
    wks = wb1[s_name]
    starting_row = 1 + read_row_num(s_name)
    starting_col = 1

    for row, item in enumerate(data):
        for col, val in enumerate(item):
            wks.cell(row=row + starting_row, column=col + starting_col).value = val

    wb1.save('{}\{}'.format(path, f_name))
    wb1.close()


def drop_dub(s_index, keep):
    """
    :param s_index: index number of the sheet to be worked on
    :type s_index int
    :param keep: keep type of the drop_dublicates function
    :type keep str
    :return: drops the dublicate rows of the chosen sheet
    """
    path2 = '{}\{}'.format(path, f_name)
    df = pd.read_excel(r'{}\{}'.format(path, f_name), sheet_name=s_index, index_col=None)
    df.drop_duplicates(keep=keep, inplace=True)

    book = load_workbook(r'{}\{}'.format(path, f_name))
    writer = pd.ExcelWriter(path2, engine='openpyxl')
    writer.book = book
    df.to_excel(writer, sheet_name='final')
    writer.save()
    writer.close()


def compare(s_base, s_check, turn):
    """
    :param turn:number of times compare function is used
    :type turn int
    :param s_base: base sheet index
    :type s_base int
    :param s_check: sheet index for the check sheet
    :type s_check int
    :return:Given a base sheet, this function tries to get the blank cell information from the check sheet.
    If the check list also does not have the information then it leaves the cell blank
    """
    final_list = []
    base_sheet = wb.sheet_by_index(s_base)
    count = 0
    for i in range(base_sheet.nrows):
        list1 = []
        row_value = base_sheet.row_values(i)
        list1.extend(row_value)
        """print(count)"""
        if "" in list1:
            for x, x in enumerate(list1):
                if x == "":
                    index_none = list1.index("")
                    if check_list(s_check, i)[index_none] != "":
                        list1[index_none] = check_list(s_check, i)[index_none]
                    else:
                        pass
        final_list.append(list1)
        count += 1
    if turn > 1:
        final_list.pop(0)
    return final_list


tic = time.perf_counter()
add_new_sheet('merge', 2)
toc = time.perf_counter()
print(f"time to add new sheet: {toc - tic:0.4f} seconds")
add_data('merge', compare(0, 1, 1))
toc2 = time.perf_counter()
print(f"time to add new data1: {toc2 - toc:0.4f} seconds")
add_data('merge', compare(1, 0, 2))
toc3 = time.perf_counter()
print(f"time to add new data2: {toc3 - toc2:0.4f} seconds")
drop_dub(2, 'first')
toc4 = time.perf_counter()
print(f"time to drop dub: {toc4 - toc3:0.4f} seconds")
remove_sheet('merge')
toc5 = time.perf_counter()
print(f"time to remove sheet: {toc5 - toc4:0.4f} seconds")
print(f"total time: {toc5 - tic:0.4f} seconds")
print('finished')
